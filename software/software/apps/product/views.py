from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, FileResponse, JsonResponse
from apps.product.forms import ProductForm, ProductEntryForm, SaleForm, UploadFileForm, DateRangeForm, CustomerForm
from apps.product.models import Product, Sale, ProductEntry, Invoice, InvoiceItem, Customer
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from apps.product.serializers import ProductSerializer
from openpyxl import load_workbook
from django.utils.timezone import now
from django.db.models import Sum, F
from datetime import date, datetime
from reportlab.pdfgen import canvas
from io import BytesIO
import xlsxwriter
import openpyxl
import datetime
import csv
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.utils import timezone
from openpyxl.utils import get_column_letter
from django.db.models import Q  # Para hacer búsquedas con OR
from django.utils.text import Truncator
from django.contrib import messages
from datetime import timedelta
from reportlab.lib.pagesizes import letter
from django.utils.timezone import make_aware
from django.template.loader import get_template
from openpyxl import Workbook
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import json
import io
from reportlab.lib.pagesizes import A4
from django.core.paginator import Paginator

# Vista index
def index(request):
    return render(request, 'products/index.html')

# Vista para crear producto
def product_view(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_list')  # Mover dentro del bloque if
    else:
        form = ProductForm()
    return render(request, 'products/product_form.html', {'form': form})

# Vista para listar productos
def product_list(request):
    products = Product.objects.all().order_by('id')  # Corregir nombre del modelo
    contexto = {'products': products}
    return render(request, 'products/product_list.html', contexto)

# Vista para editar producto
#def product_edit(request, id_product):  # Corregir nombre del parámetro
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    original_entry_date = product.entry_date
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save(commit=False)
            if not product.entry_date:
                product.entry_date = original_entry_date
            product.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'product_edit.html', {'form': form})

# Vista para eliminar producto
def product_delete(request, id_product):  # Corregir nombre del parámetro
    product = get_object_or_404(Product, id=id_product)  # Usar get_object_or_404
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'products/product_delete.html', {'product': product})

# Clase para listar productos con paginación
class ProductList(ListView):
    model = Product
    template_name = 'products/product_list.html'  # Ruta del template para listar productos
    context_object_name = 'products'  # El nombre del contexto para acceder a los productos en el template
    paginate_by = 5

    def get_queryset(self):
        query = self.request.GET.get('q')
        queryset = Product.objects.all()
        
        if query:
            # Búsqueda por nombre o marca
            queryset = queryset.filter(
                Q(product_name__icontains=query) | Q(brand__icontains=query)  
            )
        
        return queryset

# Clase para crear producto, con esta clase se está sumando la cantidad  que se ingresa si el producto ya existe con un nombre igual
class ProductCreate(CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_form.html'  # Ruta del template para crear un producto
    success_url = reverse_lazy('product_list')  # Redirigir a la lista de productos después de crear uno

    def form_valid(self, form):
        # Aquí podrías agregar alguna lógica adicional si es necesario
        return super().form_valid(form)

# Clase para actualizar producto
class ProductUpdate(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_form.html'  # Usa el mismo template para crear y actualizar
    success_url = reverse_lazy('product_list')  # Redirige a la lista después de actualizar

    def form_valid(self, form):
        # Aquí podrías agregar alguna lógica adicional si es necesario
        return super().form_valid(form)

# Clase para eliminar producto
class ProductDelete(DeleteView):
    model = Product
    template_name = 'products/product_form.html'  # Template que mostrará la confirmación para eliminar
    success_url = reverse_lazy('product_list')  # Redirige a la lista después de eliminar

# Vista para registrar clientes
class CustomerCreate(CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'products/customer_form.html'
    success_url = reverse_lazy('customer_list')

# Vista para listar clientes
class CustomerList(ListView):
    model = Customer
    template_name = 'products/customer_list.html'
    context_object_name = 'customers/list'
    paginate_by = 5 # Número de clientes por página

    def get_queryset(self):       
        query = self.request.GET.get('q')
        queryset = Customer.objects.all()
        
        if query:
            # Búsqueda por nombre, ID/NIT o apellido
            queryset = queryset.filter(
                Q(firstname__icontains=query) | Q(id_or_nit__icontains=query)  | Q(lastname__icontains=query) 
            )
        
        return queryset

# Detalle de un producto
def product_detail(request, id):
    product = get_object_or_404(Product, id=id)
    return render(request, 'products/product_detail.html', {'product': product})

# API para listar y crear productos
@api_view(['GET', 'POST'])
def product_api_list(request):
    if request.method == 'GET':
        products = Product.objects.all()
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# API para ver, actualizar y eliminar un producto específico
@api_view(['GET', 'PUT', 'DELETE'])
def product_api_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'GET':
        serializer = ProductSerializer(product)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = ProductSerializer(product, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        product.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

def import_products(request):
    """Vista para importar productos desde un archivo Excel."""
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                header_mapping = {
                    'Nombre': 'product_name',
                    'Precio': 'entry_price',
                    'Cantidad': 'entered_amount',
                    'Fecha': 'entry_date',
                    'Marca': 'brand',
                }

                headers = [cell.value for cell in sheet[1]]  # Primera fila como encabezados
                mapped_headers = [header_mapping.get(header, None) for header in headers]

                if None in mapped_headers:
                    messages.error(request, "Los encabezados del archivo no son correctos.")
                    return redirect('import_products')

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(mapped_headers, row))

                    product_name = data['product_name']
                    brand = data['brand']
                    entered_amount = data['entered_amount']
                    entry_price = data['entry_price']
                    entry_date = data['entry_date'] or timezone.now().date()

                    if not product_name or not brand or not entered_amount or not entry_price:
                        messages.error(request, f"Datos incompletos en la fila: {row}")
                        continue

                    try:
                        product = Product.objects.get(product_name=product_name, brand=brand)
                    except Product.DoesNotExist:
                        product = Product.objects.create(product_name=product_name, brand=brand)

                    # Buscar si ya existe una entrada para este producto y fecha
                    existing_entry = ProductEntry.objects.filter(
                        product=product,
                        entry_date=entry_date
                    ).first()

                    if existing_entry:
                        # Actualizar la cantidad y precio promedio de entrada
                        total_amount = existing_entry.entered_amount + entered_amount
                        existing_entry.entry_price = (
                            (existing_entry.entry_price * existing_entry.entered_amount) + (entry_price * entered_amount)
                        ) / total_amount
                        existing_entry.entered_amount = total_amount
                        existing_entry.save()
                    else:
                        # Crear una nueva entrada si no existe
                        ProductEntry.objects.create(
                            product=product,
                            entered_amount=entered_amount,
                            entry_price=entry_price,
                            entry_date=entry_date
                        )

                    # Actualizar la cantidad disponible del producto
                    product.update_available_quantity()

                messages.success(request, "Productos importados correctamente.")
                return redirect('product_list')

            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
                return redirect('import_products')

        else:
            messages.error(request, "Error al cargar el archivo. Asegúrese de que el formato sea correcto.")

    else:
        form = UploadFileForm()

    return render(request, 'products/import_products.html', {'form': form})


def generate_pdf(html_string):
    pdf = BytesIO()
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("utf-8")), dest=pdf)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', content_type='text/plain')
    return pdf.getvalue()

def product_entry_report_by_date_range(request):
    form = DateRangeForm()
    product_entries = []
    total_entries = 0
    total_quantity = 0  # Inicializar el total de cantidades ingresadas

    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            # Filtrar y ordenar productos ingresados entre las fechas seleccionadas
            product_entries = ProductEntry.objects.filter(entry_date__range=[start_date, end_date]).order_by('-entry_date')
            total_entries = product_entries.count()
            total_quantity = sum(entry.entered_amount for entry in product_entries)  # Calcular el total de cantidades ingresadas

    return render(request, 'products/product_entry_report_by_date_range.html', {
        'form': form,
        'product_entries': product_entries,
        'total_entries': total_entries,
        'total_quantity': total_quantity,  # Pasar el total de cantidades ingresadas al contexto
    })

# Exportar el reporte en PDF en rango de fecha
def export_product_entry_report_range_pdf(request, start_date, end_date):
    # Filtrar productos por el rango de fechas de ingreso
    product_entries = ProductEntry.objects.filter(entry_date__range=[start_date, end_date]).order_by('-entry_date')
    total_products = product_entries.count()  # Total de productos
    total_quantity = sum([entry.entered_amount for entry in product_entries])  # Total de cantidades

    # Generar el contenido HTML para el PDF
    html_string = render_to_string('products/product_entry_report_by_date_range_pdf.html', {
        'product_entries': product_entries,  # Asegurarse de usar product_entries en lugar de products
        'start_date': start_date,
        'end_date': end_date,
        'total_products': total_products,
        'total_quantity': total_quantity,
    })

    # Generar el PDF
    pdf = generate_pdf(html_string)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ingresos_{start_date}_a_{end_date}.pdf"'
    return response

def export_product_entry_report_range_excel(request, start_date, end_date):
    # Filtrar productos por el rango de fechas de ingreso
    product_entries = ProductEntry.objects.filter(entry_date__range=[start_date, end_date]).order_by('-entry_date')
    total_products = product_entries.count()  # Total de productos
    total_quantity = sum([entry.entered_amount for entry in product_entries])  # Total de cantidades

    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Ingreso Productos Por Rango de Fecha"

    # Escribir encabezados y títulos
    ws['A1'] = 'Reporte Ingreso Productos Por Rango de Fecha'
    ws['A2'] = f'Rango de Fecha: De {start_date} hasta {end_date}'
    ws['A4'] = 'Nombre del Producto'
    ws['B4'] = 'Precio'
    ws['C4'] = 'Cantidades Ingresadas'
    ws['D4'] = 'Marca'
    ws['E4'] = 'Fecha Ingreso'

    # Escribir los datos de productos a partir de la fila 5
    row_num = 5
    for entry in product_entries:
        ws[f'A{row_num}'] = entry.product.product_name  # Acceder al nombre del producto desde el modelo Product
        ws[f'B{row_num}'] = entry.entry_price
        ws[f'C{row_num}'] = entry.entered_amount
        ws[f'D{row_num}'] = entry.product.brand  # Acceder a la marca del producto desde el modelo Product
        ws[f'E{row_num}'] = entry.entry_date.strftime('%Y-%m-%d')
        row_num += 1

    # Escribir el total de la cantidad y productos al final
    ws[f'A{row_num+1}'] = 'Total Productos:'
    ws[f'B{row_num+1}'] = total_products
    ws[f'A{row_num+2}'] = 'Total Cantidades Ingresadas:'
    ws[f'B{row_num+2}'] = total_quantity

    # Ajustar automáticamente el ancho de las columnas según el contenido
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Generar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ingresos_{start_date}_a_{end_date}.xlsx"'
    wb.save(response)

    return response

def product_entry_report_day(request):
    today = timezone.now().date()  # Fecha de hoy
    form = ProductEntryForm(initial={'entry_date': today})

    # Filtramos por las entradas de productos del día de hoy
    product_entries = ProductEntry.objects.filter(entry_date=today).order_by('-entry_date')
    total_products = product_entries.count()
    total_quantity = product_entries.aggregate(total_quantity=Sum('entered_amount'))['total_quantity'] or 0  # Calcular la cantidad total ingresada
    selected_date = today  # Establecer la fecha seleccionada como hoy por defecto

    if request.method == 'POST':
        form = ProductEntryForm(request.POST)
        if form.is_valid():
            entry_date = form.cleaned_data['entry_date']
            selected_date = entry_date  # La fecha seleccionada por el usuario
            product_entries = ProductEntry.objects.filter(entry_date=entry_date).order_by('-entry_date')
            total_products = product_entries.count()
            total_quantity = product_entries.aggregate(total_quantity=Sum('entered_amount'))['total_quantity'] or 0  # Calcular la cantidad total ingresada

    return render(request, 'products/product_entry_report_day.html', {
        'form': form,
        'product_entries': product_entries,
        'total_products': total_products,
        'total_quantity': total_quantity,  # Pasar el total de unidades a la plantilla
        'selected_date': selected_date,
    })

# Vista para registrar una nueva entrada de producto
def product_entry_create(request):
    if request.method == 'POST':
        form = ProductEntryForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data['product']
            entered_amount = form.cleaned_data['entered_amount']
            entry_price = form.cleaned_data['entry_price']
            entry_date = form.cleaned_data['entry_date']

            # Buscar si ya existe una entrada con el mismo producto y la misma fecha
            existing_entry = ProductEntry.objects.filter(product=product, entry_date=entry_date).first()

            if existing_entry:
                # Si ya existe, sumar la cantidad y actualizar el precio si es necesario
                existing_entry.entered_amount += entered_amount
                existing_entry.entry_price = entry_price  # Puedes actualizar el precio o dejar el anterior
                existing_entry.save()
                messages.success(request, f'Se ha actualizado la cantidad del producto "{product}" ingresado el {entry_date}.')
            else:
                # Si no existe, crear una nueva entrada
                form.save()
                messages.success(request, f'Se ha registrado el producto "{product}" correctamente.')

            return redirect('product_entry_report_day')  # Redirigir al reporte de productos ingresados
    else:
        form = ProductEntryForm()  # Si es GET, mostrar el formulario vacío
    
    return render(request, 'products/product_entry_form.html', {'form': form})

# Exportar a PDF el reporte diario de ingreso de productos
def generate_pdf(html_string):
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode('utf-8')), result)
    if not pdf.err:
        return result.getvalue()
    return None

# Vista ajustada para exportar el PDF y descargarlo
def export_product_entry_day_pdf(request, entry_date):
    # Filtrar las entradas de productos por la fecha seleccionada
    product_entries = ProductEntry.objects.filter(entry_date=entry_date).order_by('-entry_date')
    total_products = product_entries.count()
    total_quantity = sum([entry.entered_amount for entry in product_entries])  # Total de cantidades

    # Generar el contenido HTML para el PDF
    html_string = render_to_string('products/product_entry_report_day_pdf.html', {
        'product_entries': product_entries,  # Asegurarse de pasar product_entries en lugar de products
        'total_products': total_products,
        'total_quantity': total_quantity,
        'entry_date': entry_date,
    })

    pdf = generate_pdf(html_string)  # Usando función que genera PDF
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        # Añadir encabezado para forzar la descarga del archivo
        response['Content-Disposition'] = f'attachment; filename="reporte_ingreso_productos_{entry_date}.pdf"'
        return response
    return HttpResponse('Error al generar el PDF', status=500)

def export_product_entry_day_excel(request, entry_date):
    # Filtrar las entradas de productos por la fecha seleccionada
    product_entries = ProductEntry.objects.filter(entry_date=entry_date).order_by('-entry_date')
    total_products = product_entries.count()
    total_quantity = sum([entry.entered_amount for entry in product_entries])

    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ingreso de Productos Diario del Día"

    # Escribir encabezados y títulos
    ws['A1'] = 'Reporte de Ingreso de Productos Diario'
    ws['A2'] = f'Fecha de Generación: {entry_date}'
    ws['A4'] = 'Nombre del Producto'
    ws['B4'] = 'Precio'
    ws['C4'] = 'Cantidades Ingresadas'
    ws['D4'] = 'Marca'
    ws['E4'] = 'Fecha Ingreso'

    # Escribir los datos de productos a partir de la fila 5
    row_num = 5
    for entry in product_entries:
        ws[f'A{row_num}'] = entry.product.product_name
        ws[f'B{row_num}'] = entry.entry_price
        ws[f'C{row_num}'] = entry.entered_amount
        ws[f'D{row_num}'] = entry.product.brand
        ws[f'E{row_num}'] = entry.entry_date.strftime('%Y-%m-%d')
        row_num += 1

    # Escribir el total de productos y la cantidad al final
    ws[f'A{row_num+1}'] = 'Total Productos:'
    ws[f'B{row_num+1}'] = total_products
    ws[f'A{row_num+2}'] = 'Total Cantidades Ingresadas:'
    ws[f'B{row_num+2}'] = total_quantity

    # Ajustar automáticamente el ancho de las columnas según el contenido
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Generar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="ingresos_{entry_date}.xlsx"'
    wb.save(response)

    return response

# Función para generar PDF
def generar_pdf(request):
    productos = Product.objects.all()
    for producto in productos:
        producto.nombre_corto = Truncator(producto.product_name).chars(30)  # Limita a 30 caracteres

#esta es la función para una sección que había de reportar venta, pero que se desactivó el día 15/oct/2024
'''def report_sale(request):
    if request.method == 'POST':
        formset = SaleFormSet(request.POST)
        error_messages = []
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data:
                    product = form.cleaned_data['product']
                    quantity_sold = form.cleaned_data['quantity_sold']

                    # Validar si hay suficiente cantidad disponible
                    if quantity_sold > product.available_quantity:
                        error_messages.append(f"No hay suficiente stock para: {product.product_name}. Cantidad disponible: {product.available_quantity}.")
                    else:
                        # Reducir la cantidad disponible y guardar la venta
                        product.available_quantity -= quantity_sold
                        product.save()
                        form.save()

            if error_messages:
                for message in error_messages:
                    messages.error(request, message)
                return render(request, 'products/report_sale.html', {'formset': formset})

            return redirect('product_list')
    else:
        formset = SaleFormSet(queryset=Sale.objects.none())

    return render(request, 'products/report_sale.html', {'formset': formset})'''

# Reporte de ventas diarias, función en donde está la grilla de resultados de las ventas y donde están los botones de exportar pdf y exportar excel
def sales_report_day(request):
    today = timezone.now().date()
    sales = Sale.objects.filter(sale_date__date=today)
    total_sales_quantity = sales.aggregate(total_quantity=Sum('quantity_sold'))['total_quantity']
    total_sales_value = sales.aggregate(total_value=Sum(F('quantity_sold') * F('product__product_entries__entry_price')))['total_value']
    
    context = {
        'sales': sales,
        'total_sales_quantity': total_sales_quantity,
        'total_sales_value': total_sales_value,
        'report_date': today,
    }
    return render(request, 'products/sales_report_day.html', context)

    
# Reporte de ventas por rango de fechas, función en donde está la grilla de resultados de las ventas y donde están los botones de exportar pdf y exportar excel
def sales_report_by_date_range(request):
    form = DateRangeForm()
    sales = []
    total_sales_quantity = 0
    total_sales_value = 0

    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            # Agrupar las ventas por producto, fecha de venta, y sumar cantidades vendidas
            sales = Sale.objects.filter(sale_date__range=[start_date, end_date]).values(
                'product__product_name',
                'product__brand',
                'product__product_entries__entry_price',
                'sale_date'
            ).annotate(
                total_quantity_sold=Sum('quantity_sold'),
                total_value=Sum('quantity_sold') * F('product__product_entries__entry_price')
            ).order_by('sale_date')

            # Calcular los totales de unidades vendidas y el valor total de ventas
            total_sales_quantity = sales.aggregate(total_quantity=Sum('total_quantity_sold'))['total_quantity'] or 0
            total_sales_value = sum([sale['total_quantity_sold'] * sale['product__product_entries__entry_price'] for sale in sales])

    return render(request, 'products/sales_report_by_date_range.html', {
        'form': form,
        'sales': sales,
        'total_sales_quantity': total_sales_quantity,
        'total_sales_value': total_sales_value,
    })

# Función para exportar reporte de ventas por rango de fechas en formato pdf 
def export_to_pdf(request, start_date, end_date):
    sales = Sale.objects.filter(sale_date__range=[start_date, end_date]).values(
        'product__product_name',
        'product__brand',
        'product__product_entries__entry_price',
        'sale_date'
    ).annotate(
        total_quantity_sold=Sum('quantity_sold'),
        total_value=Sum('quantity_sold') * F('product__product_entries__entry_price')
    ).order_by('sale_date')

    total_sales_quantity = sales.aggregate(total_quantity=Sum('total_quantity_sold'))['total_quantity'] or 0
    total_sales_value = sum([sale['total_quantity_sold'] * sale['product__product_entries__entry_price'] for sale in sales])

    template = get_template('products/sales_report_by_date_range_pdf.html')
    context = {
        'sales': sales,
        'start_date': start_date,
        'end_date': end_date,
        'total_sales_quantity': total_sales_quantity,
        'total_sales_value': total_sales_value,
    }
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.pdf"'
        return response
    return HttpResponse('Error al generar el PDF', status=400)

# Función para exportar reporte de ventas por rango de fechas en formato excel
def export_to_excel(request, start_date, end_date):
    sales = Sale.objects.filter(sale_date__range=[start_date, end_date]).values(
        'product__product_name', 'product__product_entries__entry_price', 'product__brand'
    ).annotate(total_quantity_sold=Sum('quantity_sold'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas por Rango de Fechas"
    ws['A1'] = 'Reporte de Ventas por Rango de Fechas'
    ws['A2'] = f'De {start_date} hasta {end_date}'
    ws['A4'] = 'Nombre del Producto'
    ws['B4'] = 'Precio Venta'
    ws['C4'] = 'Cantidades Vendidas'
    ws['D4'] = 'Marca'

    row_num = 5
    for sale in sales:
        ws.append([
            sale['product__product_name'],
            sale['product__product_entries__entry_price'],
            sale['total_quantity_sold'],
            sale['product__brand'],
        ])
        row_num += 1

    total_sales_quantity = sales.aggregate(total_quantity=Sum('total_quantity_sold'))['total_quantity']
    total_sales_value = sales.aggregate(total_value=Sum(F('total_quantity_sold') * F('product__product_entries__entry_price')))['total_value']

    ws[f'A{row_num+1}'] = 'Total de unidades vendidas:'
    ws[f'B{row_num+1}'] = total_sales_quantity
    ws[f'A{row_num+2}'] = 'Valor total de ventas:'
    ws[f'B{row_num+2}'] = total_sales_value

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'
    wb.save(response)
    return response

# función para exportar el reporte diario de ventas en formato pdf
def sales_report_day_pdf(request):
    today = timezone.now().date()
    sales = (
        Sale.objects.filter(sale_date__date=today)
        .values('product__product_name', 'product__product_entries__entry_price', 'product__brand', 'sale_date')  # Añadir 'sale_date'
        .annotate(total_quantity_sold=Sum('quantity_sold'))
    )
    total_sales_quantity = sales.aggregate(total_quantity=Sum('total_quantity_sold'))['total_quantity']
    total_sales_value = sum([sale['total_quantity_sold'] * sale['product__product_entries__entry_price'] for sale in sales])
    context = {
        'sales': sales,
        'total_sales_quantity': total_sales_quantity,
        'total_sales_value': total_sales_value,
        'report_date': today,
    }
    html_string = render_to_string('products/sales_report_day_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{today}.pdf"'
    pisa_status = pisa.CreatePDF(html_string, dest=response)

    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF', status=400)
    return response

# función para exportar el reporte diario de ventas en formato excel
def sales_report_day_excel(request):
    today = timezone.now().date()
    sales = (
        Sale.objects.filter(sale_date__date=today)
        .values('product__product_name', 'product__product_entries__entry_price', 'product__brand', 'sale_date')
        .annotate(total_quantity_sold=Sum('quantity_sold'))
    )
    total_sales_quantity = sales.aggregate(total_quantity=Sum('total_quantity_sold'))['total_quantity']
    total_sales_value = sum([sale['total_quantity_sold'] * sale['product__product_entries__entry_price'] for sale in sales])

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f'Reporte de Ventas Diario del día {today}'

    worksheet['A1'] = 'Reporte de Ventas Diario'
    worksheet['A2'] = f'Fecha de Generación: {today}'

    worksheet['A4'] = 'Nombre del Producto'
    worksheet['B4'] = 'Precio Venta'
    worksheet['C4'] = 'Cantidades Vendidas'
    worksheet['D4'] = 'Marca'
    worksheet['E4'] = 'Fecha de Venta'

    row_num = 5
    for sale in sales:
        sale_date = sale['sale_date'].astimezone(timezone.utc).replace(tzinfo=None)  # Eliminar la información de zona horaria de 'sale_date'
        worksheet.append([
            sale['product__product_name'],
            sale['product__product_entries__entry_price'],
            sale['total_quantity_sold'],
            sale['product__brand'],
            sale_date
        ])
        row_num += 1

    worksheet[f'A{row_num + 1}'] = 'Total de unidades vendidas:'
    worksheet[f'B{row_num + 1}'] = total_sales_quantity
    worksheet[f'A{row_num + 2}'] = 'Valor total de ventas:'
    worksheet[f'B{row_num + 2}'] = total_sales_value

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{today}.xlsx"'
    workbook.save(response)
    return response

@require_POST
def add_product_entry(request):
    try:
        product_id = request.POST.get('product_id')
        amount_entered = int(request.POST.get('entered_amount'))
        # Obtener el producto por su ID
        product = Product.objects.get(id=product_id)
        # Sumar el nuevo ingreso a 'entered_amount'
        product.entered_amount += amount_entered
        # Sincronizar 'available_quantity' con el nuevo valor total de 'entered_amount'
        product.available_quantity = product.entered_amount
        # Guardar los cambios en la base de datos
        product.save()
        return JsonResponse({
            'success': True,
            'message': f'{amount_entered} unidades ingresadas exitosamente a {product.product_name}.'
        })
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Producto no encontrado.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

'''Con esta función se confirman los cambios de las ventas y al mismo tiempo se reflejan los cambios en la vista de reporte de ventas diarias'''
@require_POST
# Vista para confirmar una venta
def confirm_sale(request):
    sale_products = request.POST.getlist('products[]')  # Obtener la lista de productos
    sale_quantities = request.POST.getlist('quantities[]')  # Obtener la lista de cantidades vendidas
    errors = []
    sold_products = []

    for i, product_id in enumerate(sale_products):
        try:
            product = Product.objects.get(id=product_id)
            quantity_sold = int(sale_quantities[i])

            if product.available_quantity >= quantity_sold:
                # Revisar si ya existe una venta para este producto en la misma sesión
                existing_sale = Sale.objects.filter(product=product, sale_date__date=timezone.now().date()).first()

                if existing_sale:
                    # Actualizar la venta existente sumando la cantidad
                    existing_sale.quantity_sold += quantity_sold
                    existing_sale.save()
                else:
                    # Registrar una nueva venta
                    sale = Sale.objects.create(
                        product=product,
                        quantity_sold=quantity_sold,
                        sale_date=timezone.now()
                    )
                    print("***sale****",sale)
                # Actualizar el stock del producto
                product.available_quantity -= quantity_sold
                product.save()

                sold_products.append(product_id)

            else:
                errors.append(f"No hay suficiente stock para {product.product_name}. Disponible: {product.available_quantity}.")
     
        except Product.DoesNotExist:
            errors.append(f"Producto con ID {product_id} no encontrado.")

        return JsonResponse({
            'success': len(errors) == 0,
            'errors': errors,
            'sold_products': sold_products
        })